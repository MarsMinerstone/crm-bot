from openpyxl import Workbook
from bot import BotDB


def define_callback_data(data: str) -> dict:
    out = {}
    a = data.split("&")
    for i in a:
        b = i.split(":", 1)
        out[b[0]] = b[1]

    return out


# def xlsx_referrer_generator(referrers_list) -> str:
#     wb = Workbook()
#     ws = wb.active

#     for i, referrer in enumerate(referrers_list):

#         ws[f'A{i+1}'], ws[f'B{i+1}'] = referrer[0], referrer[1]
    
#     wb.save("data/referrals.xlsx")
#     return("data/referrals.xlsx")
    

# def send_to_admin(user, summ, address):
#     print("d")
